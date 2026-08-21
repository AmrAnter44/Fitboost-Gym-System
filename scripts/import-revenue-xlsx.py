#!/usr/bin/env python3
# scripts/import-revenue-xlsx.py
# -------------------------------------------------------------
# استبدال إيرادات الاشتراكات في قاعدة البيانات ببيانات شيت
# "الايردات التفصيليه جيم.xlsx" — للأيام اللي الشيت بيغطيها بس.
#
# المدى: من ٩ يناير ٢٠٢٦ لحد ٣ مايو ٢٠٢٦ (بتوقيت القاهرة المحلي).
#   - أي يوم بعد كده (٤ مايو وطالع) مابيتلمسش خالص، لأن النظام
#     من التاريخ ده ماشي بشغل حقيقي مسجّل عليه.
#
# اللي بيتمسح جوّه المدى: إيصالات الاشتراكات بس
#   (Member / membershipRenewal / ترقية باكدج)
# اللي بيفضل زي ما هو: الزائر اليومي و PT و Other وأي نوع تاني.
#
# ⚠️ كمان بيصلّح باگ في الاستيراد القديم: كان بيخزّن createdAt نص
#    بدل رقم. SQLite بيرتّب النص بعد أي رقم، فأي فلتر بتاريخ كان
#    بيتجاهل الصفوف دي — يعني التقارير كانت بتطلع أرقام غلط.
#    الخطوة دي بتحوّل الصيغة بس، من غير ما تغيّر أي قيمة.
#
# ملاحظات:
#   - تواريخ الشيت محلية (القاهرة). بتتحول لـ UTC epoch ms زي ما
#     النظام بيخزن بالظبط، عشان التقارير تحسبها صح.
#   - الحركة "ارتجاع" بتتسجل بمبلغ سالب عشان صافي الإيراد يظبط.
#   - الحركة "ملغاه" بتتسجل بمبلغ صفر و isCancelled = true.
#   - أرقام فواتير الشيت مش فريدة (التقسيط بياخد نفس رقم الفاتورة)،
#     فالـ receiptNumber بياخد أرقام جديدة متسلسلة، ورقم الفاتورة
#     الأصلي بيتحفظ جوّه itemDetails في invoiceNumber.
#   - كل الصفوف المضافة id بتاعها بيبدأ بـ 'cxls' عشان تتعرف عليها
#     بسهولة لو حبيت ترجّعها.
#
# التشغيل:
#   XLSX_PASSWORD=xxxx python3 scripts/import-revenue-xlsx.py --dry-run
#   XLSX_PASSWORD=xxxx python3 scripts/import-revenue-xlsx.py
#
# المتطلبات:  pip3 install msoffcrypto-tool openpyxl
# -------------------------------------------------------------

import os
import io
import re
import sys
import json
import sqlite3
import secrets
import datetime
from zoneinfo import ZoneInfo

import msoffcrypto
import openpyxl

# ⚙️ إعدادات
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'prisma', 'الايردات التفصيليه جيم.xlsx')
DB = os.environ.get('GYM_DB', os.path.join(ROOT, 'prisma', 'gym.db'))
SHEET = 'gym-payments'

CAIRO = ZoneInfo('Africa/Cairo')
CUTOFF_LOCAL = datetime.datetime(2026, 5, 4, 0, 0, 0)   # الصفوف قبل ده بس
MEMBERSHIP_TYPES = ('Member', 'membershipRenewal', 'ترقية باكدج')
ID_PREFIX = 'cxls'
IMPORT_TAG = 'الايردات التفصيليه جيم.xlsx'

DRY_RUN = '--dry-run' in sys.argv

# أعمدة الشيت (1-based)
C_DATE, C_MOVE, C_MEMNO, C_REP, C_NAME = 2, 3, 4, 5, 6
C_PHONE, C_INVOICE, C_PLAN, C_OFFER = 8, 9, 10, 11
C_START, C_END, C_FIRST = 12, 13, 14
C_PRICE, C_REMAIN, C_BY, C_ACCOUNT, C_PAID = 17, 18, 19, 20, 21

PAY_MAP = {
    'نقدى': 'cash', 'نقدي': 'cash', 'كاش': 'cash',
    'انستا باي': 'instapay', 'إنستا باي': 'instapay',
    'محفظه': 'wallet', 'محفظة': 'wallet',
    'فيزا': 'visa',
}


def load_sheet():
    """بيفك تشفير الملف ويرجّع الـ worksheet."""
    pw = os.environ.get('XLSX_PASSWORD')
    if not pw:
        sys.exit('لازم تحط باسورد الملف في XLSX_PASSWORD')
    with open(XLSX, 'rb') as fh:
        office = msoffcrypto.OfficeFile(fh)
        office.load_key(password=pw)
        buf = io.BytesIO()
        office.decrypt(buf)
    wb = openpyxl.load_workbook(buf, data_only=True)
    return wb[SHEET]


def to_utc_ms(local_dt):
    """وقت القاهرة المحلي → UTC epoch بالمللي ثانية (زي ما النظام بيخزن)."""
    return int(local_dt.replace(tzinfo=CAIRO).timestamp() * 1000)


def iso_utc(local_dt):
    if local_dt is None:
        return None
    return local_dt.replace(tzinfo=CAIRO).astimezone(datetime.timezone.utc) \
                   .strftime('%Y-%m-%dT%H:%M:%S.000Z')


def new_id(i):
    return f'{ID_PREFIX}{i:06d}{secrets.token_hex(8)}'


def effective_local(raw):
    """
    التاريخ المخزّن في الداتابيز → datetime محلي.
    الصفوف القديمة مخزّنة بصيغتين: INTEGER (UTC epoch ms) و TEXT (ISO-8601 بتوقيت UTC،
    بلاحقة Z). الاتنين UTC — التحويل لتوقيت القاهرة بيتعمل هنا.
    """
    if isinstance(raw, int):
        return datetime.datetime.fromtimestamp(raw / 1000, datetime.timezone.utc) \
                                .astimezone(CAIRO).replace(tzinfo=None)
    # ⚠️ الـ TEXT دي UTC حقيقي (بلاحقة Z) — مش وقت محلي. قراءتها كمحلي كانت
    #    بترجّع كل إيصال 2–3 ساعات لورا، واللي وقته بعد نص الليل كان بيقع في اليوم اللي فات.
    txt = str(raw).replace('T', ' ').replace('Z', '').strip()
    naive_utc = datetime.datetime.strptime(txt[:19], '%Y-%m-%d %H:%M:%S')
    return naive_utc.replace(tzinfo=datetime.timezone.utc) \
                    .astimezone(CAIRO).replace(tzinfo=None)


def build_rows(ws, member_ids):
    """بيقرأ الشيت ويحوّله لصفوف جاهزة للإدخال."""
    rows, unmatched = [], []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, C_DATE).value
        if not isinstance(d, datetime.datetime) or d >= CUTOFF_LOCAL:
            continue

        move = ws.cell(r, C_MOVE).value
        paid = ws.cell(r, C_PAID).value or 0
        memno = ws.cell(r, C_MEMNO).value
        memno = str(int(memno)) if isinstance(memno, (int, float)) else (str(memno).strip() if memno else None)
        member_id = member_ids.get(memno)
        if memno and not member_id:
            unmatched.append((memno, ws.cell(r, C_NAME).value))

        start, end = ws.cell(r, C_START).value, ws.cell(r, C_END).value
        days = (end - start).days if isinstance(start, datetime.datetime) and isinstance(end, datetime.datetime) else None

        by = (ws.cell(r, C_BY).value or '').split('@')[0].strip() or None
        method = PAY_MAP.get((ws.cell(r, C_ACCOUNT).value or '').strip(), 'cash')

        details = {
            'memberNumber': memno,
            'memberName': ws.cell(r, C_NAME).value,
            'phone': ws.cell(r, C_PHONE).value,
            'subscriptionPrice': ws.cell(r, C_PRICE).value or 0,
            'paidAmount': paid,
            'remainingAmount': ws.cell(r, C_REMAIN).value or 0,
            'staffName': by,
            'salesPersonName': ws.cell(r, C_REP).value,
            'offerName': ws.cell(r, C_OFFER).value,
            'startDate': iso_utc(start) if isinstance(start, datetime.datetime) else None,
            'expiryDate': iso_utc(end) if isinstance(end, datetime.datetime) else None,
            'subscriptionDays': days,
            'paymentPlan': ws.cell(r, C_PLAN).value,
            'invoiceNumber': ws.cell(r, C_INVOICE).value,
            'movementType': move,
            'importSource': IMPORT_TAG,
        }

        rows.append({
            'local': d,
            'createdAt': to_utc_ms(d),
            'type': 'Member' if ws.cell(r, C_FIRST).value == 'نعم' else 'membershipRenewal',
            'amount': float(paid),
            'itemDetails': json.dumps(details, ensure_ascii=False),
            'paymentMethod': json.dumps({'methods': [{'method': method, 'amount': float(paid)}]}, ensure_ascii=False),
            'staffName': by,
            'memberId': member_id,
            'isCancelled': 1 if move == 'ملغاه' else 0,
            'cancelReason': 'ملغاه — من شيت الإيرادات' if move == 'ملغاه' else None,
        })

    rows.sort(key=lambda x: x['local'])
    return rows, unmatched


def main():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    member_ids = {str(n): i for n, i in cur.execute(
        'SELECT memberNumber, id FROM Member WHERE memberNumber IS NOT NULL')}

    ws = load_sheet()
    rows, unmatched = build_rows(ws, member_ids)

    # الإيصالات المرشحة للمسح: اشتراكات جوّه المدى (بالتوقيت المحلي)
    doomed = [dict(x) for x in cur.execute(
        f"SELECT id, receiptNumber, amount, createdAt FROM Receipt "
        f"WHERE type IN ({','.join('?' * len(MEMBERSHIP_TYPES))})", MEMBERSHIP_TYPES)]
    doomed = [x for x in doomed if effective_local(x['createdAt']) < CUTOFF_LOCAL]

    max_rn = cur.execute('SELECT MAX(receiptNumber) FROM Receipt').fetchone()[0] or 0

    # الإيصالات الباقية المخزّنة نص — بيتحوّلوا لرقم بس، قيمهم ما بتتغيرش
    doomed_ids = {x['id'] for x in doomed}
    to_fix = [(to_utc_ms(effective_local(x['createdAt'])), x['id'])
              for x in cur.execute("SELECT id, createdAt FROM Receipt WHERE typeof(createdAt) = 'text'")
              if x['id'] not in doomed_ids]

    print(f'قاعدة البيانات: {DB}')
    print(f'المدى: 2026-01-09 → 2026-05-03 (بتوقيت القاهرة)')
    print(f'هيتمسح : {len(doomed):5d} إيصال اشتراك  بإجمالي {sum(x["amount"] for x in doomed):12,.0f}')
    print(f'هيتضاف : {len(rows):5d} صف من الشيت    بإجمالي {sum(x["amount"] for x in rows):12,.0f}')
    print(f'هيتصلح: {len(to_fix):5d} إيصال تاريخه متخزّن نص (تحويل صيغة بس)')
    print(f'أرقام الإيصالات الجديدة: {max_rn + 1} → {max_rn + len(rows)}')
    if unmatched:
        print(f'⚠️  {len(unmatched)} صف رقم عضويته مش موجود في الداتابيز (هيتسجل من غير ربط بعضو): {unmatched}')

    if DRY_RUN:
        print('\n--dry-run — مفيش أي تعديل اتعمل.')
        conn.close()
        return

    try:
        cur.execute('BEGIN')
        cur.executemany('DELETE FROM Receipt WHERE id = ?', [(x['id'],) for x in doomed])
        cur.executemany(
            'INSERT INTO Receipt (id, receiptNumber, type, amount, itemDetails, paymentMethod,'
            ' createdAt, staffName, memberId, isCancelled, cancelledAt, cancelledBy, cancelReason)'
            ' VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
            [(new_id(i), max_rn + 1 + i, r['type'], r['amount'], r['itemDetails'], r['paymentMethod'],
              r['createdAt'], r['staffName'], r['memberId'], r['isCancelled'],
              r['createdAt'] if r['isCancelled'] else None,
              'استيراد' if r['isCancelled'] else None, r['cancelReason'])
             for i, r in enumerate(rows)])
        cur.executemany('UPDATE Receipt SET createdAt = ? WHERE id = ?', to_fix)
        cur.execute('UPDATE ReceiptCounter SET current = ? WHERE current < ?',
                    (max_rn + len(rows), max_rn + len(rows)))
        conn.commit()
        print('\n✅ تم.')
    except Exception as e:
        conn.rollback()
        sys.exit(f'\n❌ فشل — اترجع كل حاجة زي ما كانت: {e}')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
