#!/usr/bin/env python3
# scripts/import-old-gym-sheets.py
# -------------------------------------------------------------
# استيراد داتا الجيم القديم (شيتات إكسل في ~/Desktop/shets) داخل
# قاعدة بيانات FitBoost فاضية (prisma/gym-new.db).
#
# اللي بيتستورد:
#   1. شيت التأهيل الرئيسي.xlsx → شيتات "تامر 8" و"محمود8" (أغسطس 2026 فقط)
#        → اشتراك More لكل عميل باسم المدرب (تامر / محمود)
#        → MoreSession لكل يوم حضور (علامة في خانة اليوم)
#        → العمود اللي جنب EX = عدد الجلسات المتبقية
#        → الباقات: 6 = أسبوع، 10 = أسبوعين، 12 = 3 أسابيع، 16 = شهر،
#          "كامل" = 30 جلسة في 30 يوم. المشتراه = أقرب باقة فوق (المتبقي + الحضور)
#   2. inv.xlsx → شيت "black list" → BannedMember
#   3. inv.xlsx → شيتات "ك بكر" و"ك عبد الله 7/8" → اشتراكات More
#        باسم الكوتش (بكر / عبد الله)، الكود والمدة في الملاحظات
#   4. Staff مدرب لكل من: تامر، محمود، بكر، عبد الله
#
# اللي مش بيتستورد (بقرار صاحب الجيم): شيتات التجديد/التارجت،
# وشيت الدعوات inv (محتاج أعضاء مسجّلين).
#
# التشغيل:
#   python3 scripts/import-old-gym-sheets.py --dry-run
#   python3 scripts/import-old-gym-sheets.py
#   GYM_DB=/path/to/other.db SHEETS_DIR=/path/to/sheets python3 scripts/import-old-gym-sheets.py
#
# كل الصفوف المضافة id بتاعها بيبدأ بـ 'cimp' عشان تتعرف عليها بسهولة.
# المتطلبات: pip3 install openpyxl
# -------------------------------------------------------------
import os
import re
import sys
import sqlite3
import secrets
import datetime
from zoneinfo import ZoneInfo

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.environ.get('GYM_DB', os.path.join(ROOT, 'prisma', 'gym-new.db'))
SHEETS_DIR = os.environ.get('SHEETS_DIR', os.path.expanduser('~/Desktop/shets'))
REHAB_XLSX = os.path.join(SHEETS_DIR, 'شيت التأهيل الرئيسي.xlsx')
INV_XLSX = os.path.join(SHEETS_DIR, 'inv.xlsx')
DRY_RUN = '--dry-run' in sys.argv

CAIRO = ZoneInfo('Africa/Cairo')
NOW = datetime.datetime.now(CAIRO)
NOW_MS = int(NOW.timestamp() * 1000)
IMPORT_TAG = 'استيراد من الشيت القديم'
DEFAULT_YEAR = 2026

# شهر شيت التأهيل المستورد
REHAB_SHEETS = [('تامر 8', 'تامر'), ('محمود8', 'محمود')]
REHAB_YEAR, REHAB_MONTH = 2026, 8
FULL_PACKAGE_SESSIONS = 30      # "كامل" = 30 جلسة في 30 يوم
# باقات التأهيل: عدد الجلسات → مدة الباقة بالأيام
REHAB_PACKAGES = [(6, 7), (10, 14), (12, 21), (16, 30), (30, 30)]


def rehab_package(total):
    """أقرب باقة تغطي (المتبقي + الحضور) → (جلسات الباقة، مدتها بالأيام)."""
    for sessions, days in REHAB_PACKAGES:
        if total <= sessions:
            return sessions, days
    return REHAB_PACKAGES[-1]
SESSIONS_PER_MONTH_KIDS = 12    # شيتات الكوتشات: 3 حصص/أسبوع تقريباً

COACH_SHEETS = [('ك بكر', 'بكر'), ('ك عبد الله 7', 'عبد الله'), ('ك عبد الله 8', 'عبد الله')]
STAFF = ['تامر', 'محمود', 'بكر', 'عبد الله']

ALNUM = 'abcdefghijklmnopqrstuvwxyz0123456789'


def cuid():
    return 'cimp' + ''.join(secrets.choice(ALNUM) for _ in range(21))


def ms(dt_local):
    """datetime محلي (القاهرة) → epoch ms زي ما Prisma بيخزن."""
    if dt_local.tzinfo is None:
        dt_local = dt_local.replace(tzinfo=CAIRO)
    return int(dt_local.timestamp() * 1000)


def clean(s):
    return re.sub(r'\s+', ' ', str(s)).strip() if s is not None else ''


def norm_name(s):
    s = clean(s)
    s = re.sub('[أإآ]', 'ا', s).replace('ى', 'ي').replace('ة', 'ه')
    return re.sub(r'\s+', '', s)


def parse_date(v, default_year=DEFAULT_YEAR):
    """يقبل datetime أو نص زي 13\\9 أو 1\\9\\2026 أو 25/8/2026 أو 16/8\\2026."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.replace(tzinfo=None)
    s = clean(v)
    if not s:
        return None
    parts = [p for p in re.split(r'[\\/\-.]+', s) if p.strip()]
    if len(parts) < 2:
        return None
    try:
        d, m = int(parts[0]), int(parts[1])
        y = int(parts[2]) if len(parts) >= 3 else default_year
        if y < 100:
            y += 2000
    except ValueError:
        return None
    if not (1 <= m <= 12):
        return None
    # تواريخ غلط زي 31\11 أو 30\2 → آخر يوم في الشهر
    last = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1)).day
    d = max(1, min(d, last))
    return datetime.datetime(y, m, d)


def parse_sessions(v):
    """العمود اللي جنب EX: رقم = جلسات متبقية، 16+16 = 32، 'كامل' = باقة كاملة."""
    if v is None:
        return None
    s = clean(v)
    if not s:
        return None
    if 'كامل' in s:
        return 'full'
    nums = re.findall(r'\d+', s)
    if not nums:
        return None
    return sum(int(n) for n in nums)


def parse_duration_months(v):
    s = clean(v)
    if not s:
        return None
    n = re.findall(r'\d+', s)
    if n:
        return int(n[0])
    if 'شهر' in s:
        return 1
    return None


def add_months(dt, months):
    m = dt.month - 1 + months
    y = dt.year + m // 12
    m = m % 12 + 1
    last = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1)).day
    return dt.replace(year=y, month=m, day=min(dt.day, last))


# ---------------- قراءة الشيتات ----------------

def read_rehab():
    """→ list of dict(name, coach, expiry, remaining, full, days[])"""
    wb = openpyxl.load_workbook(REHAB_XLSX, data_only=True)
    out, skipped = [], []
    for sheet, coach in REHAB_SHEETS:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr = rows[0]
        day1 = [i for i, v in enumerate(hdr) if v == 1][0]
        for r in rows[1:]:
            name = clean(r[0])
            if not name or 'NAME' in name.upper():
                continue
            expiry = parse_date(r[1])
            if expiry is None:
                skipped.append((sheet, name, 'بدون تاريخ EX'))
                continue
            sess = parse_sessions(r[2])
            days = [i + 1 for i in range(31)
                    if day1 + i < len(r) and clean(r[day1 + i]) not in ('', '-')]
            out.append(dict(sheet=sheet, name=name, coach=coach, expiry=expiry,
                            sessions=sess, days=days))
    return out, skipped


def read_blacklist():
    wb = openpyxl.load_workbook(INV_XLSX, data_only=True)
    ws = wb['black list']
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        name = clean(r[2])
        if not name:
            continue
        phone = re.sub(r'\D', '', clean(r[1]))
        out.append(dict(name=name, phone=phone or None, status=clean(r[0]),
                        expiry=parse_date(r[3])))
    return out


def read_coach_sheets():
    """→ list of dict(name, code, coach, expiry, months, note, sheet) بعد إزالة التكرار"""
    wb = openpyxl.load_workbook(INV_XLSX, data_only=True)
    rows_all, skipped = [], []
    for sheet, coach in COACH_SHEETS:
        ws = wb[sheet]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            name = clean(r[0])
            if not name:
                continue
            code = clean(r[1]) if len(r) > 1 else ''
            code = re.sub(r'\.0$', '', code)
            expiry = parse_date(r[2]) if len(r) > 2 else None
            months = parse_duration_months(r[3]) if len(r) > 3 else None
            extra = ' '.join(clean(v) for v in r[4:] if clean(v) and not re.fullmatch(r'\d+', clean(v)))
            if expiry is None:
                skipped.append((sheet, name, code, 'بدون تاريخ EX'))
                continue
            rows_all.append(dict(sheet=sheet, name=name, code=code, coach=coach,
                                 expiry=expiry, months=months or 1, note=extra))
    # إزالة التكرار: نفس الكوتش + نفس الكود → نسيب أحدث EX.
    # الصف اللي من غير كود بيتلحق بصف بنفس الاسم لو ليه كود (نفس الطفل اتسجل كوده بعدين).
    best = {}
    for row in rows_all:
        key = (row['coach'], ('code', row['code']) if row['code'] else ('name', norm_name(row['name'])))
        if key not in best or row['expiry'] > best[key]['expiry']:
            best[key] = row
    drop = set()
    for key, row in list(best.items()):
        if key[1][0] != 'name':
            continue
        twin = next((k for k, r in best.items()
                     if k[0] == key[0] and k[1][0] == 'code' and norm_name(r['name']) == key[1][1]), None)
        if twin:
            if row['expiry'] > best[twin]['expiry']:
                best[twin] = dict(best[twin], expiry=row['expiry'], months=row['months'], sheet=row['sheet'])
            drop.add(key)
    return [r for k, r in best.items() if k not in drop], skipped, len(rows_all)


# ---------------- الكتابة في الداتابيز ----------------

def insert_staff(cur):
    ids = {}
    for i, name in enumerate(STAFF, start=1):
        existing = cur.execute('SELECT id FROM Staff WHERE name = ?', (name,)).fetchone()
        if existing:
            ids[name] = existing[0]
            continue
        sid = cuid()
        cur.execute(
            'INSERT INTO Staff (id, staffCode, name, position, notes, isActive, createdAt, updatedAt) '
            'VALUES (?,?,?,?,?,?,?,?)',
            (sid, f's90000000{i}', name, 'مدرب', IMPORT_TAG, 1, NOW_MS, NOW_MS))
        ids[name] = sid
    return ids


def next_more_number(cur):
    row = cur.execute('SELECT COALESCE(MAX(moreNumber), 0) FROM More').fetchone()
    return row[0] + 1


def insert_more(cur, *, client, coach, start, expiry, purchased, remaining, notes):
    n = next_more_number(cur)
    cur.execute(
        'INSERT INTO More (moreNumber, clientName, phone, memberId, sessionsPurchased, sessionsRemaining, '
        'coachName, coachUserId, pricePerSession, totalAmount, startDate, expiryDate, remainingAmount, '
        'notes, isActive, createdAt, updatedAt) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (n, client, '', None, purchased, remaining, coach, None, 0.0, 0.0,
         ms(start), ms(expiry.replace(hour=23, minute=59, second=59)), 0.0, notes,
         1 if expiry.replace(tzinfo=CAIRO) >= NOW else 0, NOW_MS, NOW_MS))
    return n


def insert_session(cur, more_number, client, coach, when):
    cur.execute(
        'INSERT INTO MoreSession (id, moreNumber, clientName, coachName, sessionDate, attended, attendedAt, '
        'attendedBy, notes, createdAt, isFreeSession, memberId, collectedInExpenseId) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (cuid(), more_number, client, coach, ms(when), 1, ms(when), IMPORT_TAG, None, ms(when), 0, None, None))


def main():
    print(f'🗄️  الداتابيز: {DB}')
    print(f'📂 الشيتات:   {SHEETS_DIR}')
    if DRY_RUN:
        print('🧪 DRY RUN — مفيش حاجة هتتكتب\n')
    if not os.path.exists(DB):
        sys.exit(f'❌ الداتابيز مش موجودة: {DB}\n   اعملها الأول: DATABASE_URL="file:./gym-new.db" npx prisma db push --skip-generate')

    rehab, rehab_skipped = read_rehab()
    blacklist = read_blacklist()
    kids, kids_skipped, kids_raw = read_coach_sheets()

    con = sqlite3.connect(DB)
    cur = con.cursor()
    try:
        staff_ids = insert_staff(cur)
        print(f'👤 Staff مدرب: {", ".join(staff_ids)}')

        # 1) التأهيل → More + MoreSession
        sessions_total = 0
        active = 0
        for r in rehab:
            att = len(r['days'])
            if r['sessions'] == 'full':
                purchased, days = FULL_PACKAGE_SESSIONS, 30
                remaining = max(purchased - att, 0)
                sess_note = 'كامل (30 جلسة / 30 يوم)'
            elif r['sessions'] is None:
                purchased, days = rehab_package(att)
                remaining = purchased - att
                sess_note = 'بدون عدد جلسات في الشيت'
            else:
                remaining = r['sessions']
                purchased, days = rehab_package(remaining + att)
                sess_note = f'باقة {purchased} جلسة — متبقي في الشيت: {remaining}'
            start = r['expiry'] - datetime.timedelta(days=days)
            notes = f'تأهيل — {IMPORT_TAG} ({r["sheet"]}) — {sess_note} — حضور أغسطس: {att} يوم'
            n = insert_more(cur, client=r['name'], coach=r['coach'], start=start, expiry=r['expiry'],
                            purchased=purchased, remaining=remaining, notes=notes)
            if r['expiry'].replace(tzinfo=CAIRO) >= NOW:
                active += 1
            for d in r['days']:
                when = datetime.datetime(REHAB_YEAR, REHAB_MONTH, d, 12, 0)
                insert_session(cur, n, r['name'], r['coach'], when)
                sessions_total += 1
        print(f'🏥 التأهيل: {len(rehab)} اشتراك More ({active} نشط) + {sessions_total} جلسة حضور')
        for s in rehab_skipped:
            print('   ⚠️ متجاهل:', s)

        # 2) شيتات الكوتشات → More
        kids_active = 0
        for r in kids:
            purchased = SESSIONS_PER_MONTH_KIDS * r['months']
            start = add_months(r['expiry'], -r['months'])
            bits = [f'كوتش {r["coach"]} — {IMPORT_TAG} ({r["sheet"]})']
            if r['code']:
                bits.append(f'الكود: {r["code"]}')
            bits.append(f'المدة: {r["months"]} شهر')
            if r['note']:
                bits.append(r['note'])
            insert_more(cur, client=r['name'], coach=r['coach'], start=start, expiry=r['expiry'],
                        purchased=purchased, remaining=purchased, notes=' — '.join(bits))
            if r['expiry'].replace(tzinfo=CAIRO) >= NOW:
                kids_active += 1
        print(f'🧒 شيتات الكوتشات: {kids_raw} صف → {len(kids)} اشتراك More بعد إزالة التكرار ({kids_active} نشط)')
        for s in kids_skipped:
            print('   ⚠️ متجاهل:', s)

        # 3) Black list → BannedMember
        for b in blacklist:
            note_bits = [b['status']] if b['status'] else []
            if b['expiry']:
                note_bits.append(f'ينتهي اشتراكه {b["expiry"]:%d/%m/%Y}')
            cur.execute(
                'INSERT INTO BannedMember (id, name, phone, nationalId, reason, notes, bannedBy, createdAt, updatedAt) '
                'VALUES (?,?,?,?,?,?,?,?,?)',
                (cuid(), b['name'], b['phone'], None, 'Black list — الشيت القديم',
                 ' — '.join(note_bits) or None, IMPORT_TAG, NOW_MS, NOW_MS))
        print(f'🚫 المحظورين: {len(blacklist)}')

        if DRY_RUN:
            con.rollback()
            print('\n🧪 تم التراجع (dry run).')
        else:
            con.commit()
            print('\n✅ تم الحفظ.')
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()


if __name__ == '__main__':
    main()
